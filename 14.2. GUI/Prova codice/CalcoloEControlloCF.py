import datetime
import json
import openpyxl
from pathlib import Path


class CodiceFiscale:

    def __init__(self, nome, cognome, data_di_nascita, sesso, citta_di_nascita):
        self.nome = nome
        self.cognome = cognome
        self.data_di_nascita = data_di_nascita
        self.sesso = sesso
        self.citta_di_nascita = citta_di_nascita
        self.codice_fiscale = ""
        self.vocali = ['a', 'e', 'o', 'i', 'u']
        self.mesi_dict = {}
        self.conversione_caratteri_dispari_dict = {}
        self.conversione_caratteri_pari_dict = {}
        self.carattere_alfabetico_di_controllo_dict = {}
        self.carica_json()

    def carica_json(self):
        with open('../Database/MesiCF.json', 'r') as mesi_CF:
            self.mesi_dict = json.load(mesi_CF)
        with open('../Database/ConversioneCaratteriDispariCF.json', 'r') as conversione_caratteri_dispari:
            self. conversione_caratteri_dispari_dict = json.load(conversione_caratteri_dispari)
        with open('../Database/ConversioneCaratteriPariCF.json', 'r') as conversione_caratteri_pari:
            self.conversione_caratteri_pari_dict = json.load(conversione_caratteri_pari)
        with open('../Database/CarattereAlfabeticoDiControllo.json', 'r') as carattere_alfabetico_di_controllo:
            self.carattere_alfabetico_di_controllo_dict = json.load(carattere_alfabetico_di_controllo)

    def calcolo(self):
        self.calcolo_cognome()
        self.calcolo_nome()
        self.calcolo_data_di_nascita()
        self.calcolo_codice_catastale()
        self.calcolo_carattere_alfabetico_di_controllo()
        print(self.codice_fiscale)

    def calcolo_cognome(self):

        self.cognome = self.cognome.lower()

        for c in self.cognome:
            if not c.isalpha():
                self.cognome = self.cognome.replace(c, '')

        print(self.cognome)

        for c in self.cognome:
            if c not in self.vocali and len(self.codice_fiscale) < 3:
                self.codice_fiscale += c.upper()

        for c in self.cognome:
            if c in self.vocali and len(self.codice_fiscale) < 3:
                self.codice_fiscale += c.upper()

        while len(self.codice_fiscale) < 3:
            self.codice_fiscale += 'X'

    def calcolo_nome(self):

        self.nome = self.nome.lower()

        for c in self.nome:
            if not c.isalpha():
                self.nome = self.nome.replace(c, '')

        print(self.nome)

        consonanti = ""

        for c in self.nome:
            if c not in self.vocali:
                consonanti += c

        if len(consonanti) > 3:
            self.codice_fiscale += consonanti[0].upper()
            self.codice_fiscale += consonanti[2].upper()
            self.codice_fiscale += consonanti[3].upper()
        elif len(consonanti) == 3:
            self.codice_fiscale += consonanti.upper()
        else:
            self.codice_fiscale += consonanti.upper()

            for c in self.nome:
                if c in self.vocali and len(self.codice_fiscale) < 6:
                    self.codice_fiscale += c.upper()

            while len(self.codice_fiscale) < 6:
                self.codice_fiscale += 'X'

    def calcolo_data_di_nascita(self):

        print(self.data_di_nascita)

        anno = str(self.data_di_nascita.year)[-2:]
        mese = str(self.data_di_nascita.month)
        giorno = self.data_di_nascita.day

        for c in anno:
            self.codice_fiscale += c

        self.codice_fiscale += self.mesi_dict.get(mese)

        if self.sesso == 'M':
            self.codice_fiscale += str(giorno)
        else:
            self.codice_fiscale += str(giorno + 40)

    def calcolo_codice_catastale(self):

        file = Path('../Database', 'Comuni e codici catastali.xlsx')
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        riga = 0

        for colonna in sheet.iter_cols(max_col=1):
            for cella in colonna:
                riga += 1
                if cella.value == self.citta_di_nascita:
                    self.codice_fiscale += str(sheet.cell(row=riga, column=2).value)
                    break

        workbook.close()

    def calcolo_carattere_alfabetico_di_controllo(self):

        somma = 0

        for i in range(len(self.codice_fiscale)):
            c = self.codice_fiscale[i]
            if (i + 1) % 2 != 0:
                somma += self.conversione_caratteri_dispari_dict.get(c)
            else:
                somma += self.conversione_caratteri_pari_dict.get(c)

        resto = somma % 26

        self.codice_fiscale += self.carattere_alfabetico_di_controllo_dict.get(str(resto))


d_d_n = datetime.date(2000, 5, 11)
codice_fiscale = CodiceFiscale("Emanuele", "Frisi", d_d_n, "M", "Foggia")
codice_fiscale.calcolo()
