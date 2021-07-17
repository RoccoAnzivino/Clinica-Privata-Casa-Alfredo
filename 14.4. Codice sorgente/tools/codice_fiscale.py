import json
import openpyxl
from pathlib import Path


class CodiceFiscale:

    def __init__(self, nome, cognome, data_di_nascita, sesso, citta_di_nascita):
        """
        Costruttore della classe CodiceFiscale, nella quale viene effettuato un calcolo del codice fiscale a partire dai
        dati relativi all'utente, al fine di controllare la correttezza del codice fiscale inserito tramite confronto

        :param nome: Parametro che rappresenta il nome del paziente relativo al paziente stesso
        :type nome: str
        :param cognome: Parametro che rappresenta il cognome del paziente relativo al paziente stesso
        :type cognome: str
        :param data_di_nascita: Parametro che rappresenta la data di nascita del paziente relativa al paziente stesso
        :type data_di_nascita: date
        :param sesso: Parametro che rappresenta il sesso del paziente relativo al paziente stesso
        :type sesso: str
        :param citta_di_nascita: Parametro che rappresenta la città di nascita del paziente relativa al paziente stesso
        :type citta_di_nascita: str
        """
        self.nome = nome
        self.cognome = cognome
        self.data_di_nascita = data_di_nascita
        self.sesso = sesso
        self.citta_di_nascita = citta_di_nascita
        self.codice_fiscale = ''
        self.vocali = ['a', 'e', 'o', 'i', 'u']
        # ho cambiato il nome di queste variabili
        self.mesi = {}
        self.conversione_caratteri_dispari = {}
        self.conversione_caratteri_pari = {}
        self.carattere_alfabetico_di_controllo = {}

        self.carica_json()
        self.calcolo_codice_fiscale()

    def carica_json(self):
        """
        Funzione che legge da file 'MesiCF.json' e salva le informazioni nella relativa variabile d'istanza,
        così come anche per 'ConversìoneCaratteriDispariCF.json', 'ConversioneCaratteriPariCF.json' e
        'CarattereAlfabeticoDiControllo.json'
        """
        with open('database/MesiCF.json', 'r') as mesi_CF:
            self.mesi = json.load(mesi_CF)

        with open('database/ConversioneCaratteriDispariCF.json', 'r') as conversione_caratteri_dispari:
            self.conversione_caratteri_dispari = json.load(conversione_caratteri_dispari)

        with open('database/ConversioneCaratteriPariCF.json', 'r') as conversione_caratteri_pari:
            self.conversione_caratteri_pari = json.load(conversione_caratteri_pari)

        with open('database/CarattereAlfabeticoDiControllo.json', 'r') as carattere_alfabetico_di_controllo:
            self.carattere_alfabetico_di_controllo = json.load(carattere_alfabetico_di_controllo)

    def calcolo_codice_fiscale(self):
        """
        Funzione che richiama tutte le funzioni relative al calcolo del codice fiscale, effettuando l'effettivo calcolo
        e, infine, confronta il codice fiscale calcolato con quello inserito
        """
        self.calcolo_cognome()
        self.calcolo_nome()
        self.calcolo_data_di_nascita()
        self.calcolo_codice_catastale()
        self.calcolo_carattere_alfabetico_di_controllo()
        print(self.codice_fiscale)

    def calcolo_cognome(self):
        """
        Funzione che, a partire dalla variabile d'istanza del cognome del paziente, calcola la relativa sezione
        all'interno del codice fiscale
        """
        self.cognome = self.cognome.lower()

        for c in self.cognome:
            if not c.isalpha():
                self.cognome = self.cognome.replace(c, '')

        print(self.cognome)

        for c in self.cognome:
            if c not in self.vocali and len(self.codice_fiscale) < 3:
                self.codice_fiscale += c.upper()

        for c in self.cognome:
            if c in self.vocali and len(self.codice_fiscale) < 3:
                self.codice_fiscale += c.upper()

        while len(self.codice_fiscale) < 3:
            self.codice_fiscale += 'X'

    def calcolo_nome(self):
        """
        Funzione che, a partire dalla variabile d'istanza del nome del paziente, calcola la relativa sezione
        all'interno del codice fiscale
        """
        self.nome = self.nome.lower()

        for c in self.nome:
            if not c.isalpha():
                self.nome = self.nome.replace(c, '')

        print(self.nome)

        consonanti = ""

        for c in self.nome:
            if c not in self.vocali:
                consonanti += c

        if len(consonanti) > 3:
            self.codice_fiscale += consonanti[0].upper()
            self.codice_fiscale += consonanti[2].upper()
            self.codice_fiscale += consonanti[3].upper()
        elif len(consonanti) == 3:
            self.codice_fiscale += consonanti.upper()
        else:
            self.codice_fiscale += consonanti.upper()

            for c in self.nome:
                if c in self.vocali and len(self.codice_fiscale) < 6:
                    self.codice_fiscale += c.upper()

            while len(self.codice_fiscale) < 6:
                self.codice_fiscale += 'X'

    def calcolo_data_di_nascita(self):
        """
        Funzione che, a partire dalla variabile d'istanza della data di nascita del paziente, calcola la relativa
        sezione all'interno del codice fiscale
        """
        print(self.data_di_nascita)

        anno = str(self.data_di_nascita.year)[-2:]
        mese = str(self.data_di_nascita.month)
        giorno = self.data_di_nascita.day

        for c in anno:
            self.codice_fiscale += c

        self.codice_fiscale += self.mesi.get(mese)

        if self.sesso == 'M':
            if len(str(giorno)) > 1:
                self.codice_fiscale += str(giorno)
            else:
                self.codice_fiscale += '0' + str(giorno)
        else:
            self.codice_fiscale += str(giorno + 40)

    def calcolo_codice_catastale(self):
        """
        Funzione che legge da file 'Comuni e codici catastali.xlsx' e calcola la relativa sezione all'interno del codice
        fiscale
        """
        file = Path('database', 'Comuni e codici catastali.xlsx')
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        riga = 0

        for colonna in sheet.iter_cols(max_col=1):
            for cella in colonna:
                riga += 1
                if cella.value == self.citta_di_nascita:
                    self.codice_fiscale += str(sheet.cell(row=riga, column=2).value)
                    break

        workbook.close()

    def calcolo_carattere_alfabetico_di_controllo(self):
        """
        Funzione che calcola l'ultimo carattere del codice fiscale leggendo ogni singolo carattere di quest'ultimo,
        seguendo le indicazioni fornite dalla lettura dei file precedentemente effettuata
        """
        somma = 0

        for i in range(len(self.codice_fiscale)):
            c = self.codice_fiscale[i]
            if (i + 1) % 2 != 0:
                somma += self.conversione_caratteri_dispari.get(c)
            else:
                somma += self.conversione_caratteri_pari.get(c)

        resto = somma % 26

        self.codice_fiscale += self.carattere_alfabetico_di_controllo.get(str(resto))

    def controllo_codice_fiscale(self, codice_fiscale):
        """
        Funzione che confronta il codice fiscale passato come parametro alla funzione stessa con quello calcolato

        :param codice_fiscale: Parametro che rappresenta il codice fiscale inserito in fase di registrazione utente
        :type codice_fiscale: str
        :return: La funzione ritorna True se i due codici fiscali coincidono, altrimenti ritorna False
        :rtype: bool
        """
        if codice_fiscale == self.codice_fiscale:
            return True
        else:
            return False
