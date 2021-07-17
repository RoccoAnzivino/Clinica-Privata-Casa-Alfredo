import openpyxl
from pathlib import Path

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QWidget

from paziente.login_paziente.controller import controller_lista_pazienti


class ViewRegistrazionePaziente(QWidget):

    def __init__(self):
        """
        Costruttore della classe ViewRegistrazionePaziente, nella quale vengono creati e mostrati tutti gli oggetti
        della Graphical User Interface (GUI) relativi alla suddetta view
        """
        super().__init__()

        self.sheet = None
        self.province = []

        self.logo = QtWidgets.QLabel(self)
        self.titolo = QtWidgets.QLabel(self)
        self.scritta = QtWidgets.QLabel(self)
        self.registrati_button = QtWidgets.QPushButton(self)
        self.banner = QtWidgets.QLabel(self)
        self.torna_indietro = QtWidgets.QPushButton(self)
        self.occhiello_barrato_button = QtWidgets.QPushButton(self)
        self.occhiello_button = QtWidgets.QPushButton(self)
        self.registrazione = QtWidgets.QLabel(self)
        self.dati_personali = QtWidgets.QLabel(self)
        self.nome = QtWidgets.QLabel(self)
        self.cognome = QtWidgets.QLabel(self)
        self.sesso = QtWidgets.QLabel(self)
        self.citta_di_nascita = QtWidgets.QLabel(self)
        self.provincia = QtWidgets.QLabel(self)
        self.data_di_nascita = QtWidgets.QLabel(self)
        self.codice_fiscale = QtWidgets.QLabel(self)
        self.sesso_combo_box = QtWidgets.QComboBox(self)
        self.citta_di_nascita_combo_box = QtWidgets.QComboBox(self)
        self.provincia_combo_box = QtWidgets.QComboBox(self)
        self.nome_line_edit = QtWidgets.QLineEdit(self)
        self.cognome_line_edit = QtWidgets.QLineEdit(self)
        self.data_di_nascita_date_edit = QtWidgets.QDateEdit(self)
        self.codice_fiscale_line_edit = QtWidgets.QLineEdit(self)
        self.email_line_edit = QtWidgets.QLineEdit(self)
        self.password_line_edit = QtWidgets.QLineEdit(self)
        self.email = QtWidgets.QLabel(self)
        self.password = QtWidgets.QLabel(self)
        self.dati_account = QtWidgets.QLabel(self)
        self.setup_ui(self)

        self.controller_lista_pazienti = controller_lista_pazienti.ControllerListaPazienti(self)

    def occhiello_button_pressed(self):
        """
        Funzione che rende visibile il testo inserito nella line edit delle password nel momento in cui si preme
        l'apposito button
        """
        # non fa vedere la password digitata
        self.password_line_edit.setEchoMode(QtWidgets.QLineEdit.Password)
        self.occhiello_barrato_button.setVisible(True)
        self.occhiello_button.setVisible(False)
        # Per risolvere il bug del focus
        self.nome_line_edit.clearFocus()

    def occhiello_barrato_button_pressed(self):
        """
        Funzione che nasconde l testo inserito nella line edit delle password nel momento in cui si preme l'apposito
        button
        """
        # fa vedere la password digitata
        self.password_line_edit.setEchoMode(QtWidgets.QLineEdit.Normal)
        self.occhiello_button.setVisible(True)
        self.occhiello_barrato_button.setVisible(False)
        # Per risolvere il bug del focus
        self.nome_line_edit.clearFocus()

    def lista_comuni(self, provincia):
        """
        Funzione che mostra unicamente i comuni legati alla provincia selezionata all'interno della combobox della città
        di nascita del paziente

        :param provincia: Parametro che rappresenta la provincia selezionata dal paziente in fase di registrazione
        :type provincia: str
        """
        comuni = []
        i = 0
        n = 0

        for colonna in self.sheet.iter_cols():
            i += 1
            cella = colonna[0]
            if cella.value == provincia:
                n = i

        for colonna in self.sheet.iter_cols(n, max_col=n):
            for cella in colonna:
                if cella.value is not None and cella is not colonna[0]:
                    comuni.append(cella.value)

        self.citta_di_nascita_combo_box.clear()
        self.citta_di_nascita_combo_box.addItems(comuni)

    def setup_ui(self, registrazione_paziente):
        """
        Funzione che crea e determina le caratteristiche degli oggetti della ViewRegistrazionePaziente

        :param registrazione_paziente: Oggetto della view che rappresenta la view stessa
        :type registrazione_paziente: ViewRegistrazionePaziente
        """
        file = Path('database', 'Province e comuni.xlsx')
        workbook = openpyxl.load_workbook(file)
        self.sheet = workbook.active

        for riga in self.sheet.iter_rows(max_row=1):
            for cella in riga:
                self.province.append(cella.value)

        registrazione_paziente.setObjectName("registrazione_paziente")
        registrazione_paziente.resize(900, 600)
        registrazione_paziente.setMinimumSize(QtCore.QSize(900, 600))
        registrazione_paziente.setMaximumSize(QtCore.QSize(900, 600))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        registrazione_paziente.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        registrazione_paziente.setFont(font)
        registrazione_paziente.setMouseTracking(False)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Immagini/logo_casa_alfredo.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        registrazione_paziente.setWindowIcon(icon)
        self.logo.setGeometry(QtCore.QRect(10, 10, 35, 40))
        self.logo.setText("")
        self.logo.setPixmap(QtGui.QPixmap("Immagini/logo_casa_alfredo.png"))
        self.logo.setScaledContents(True)
        self.logo.setObjectName("logo")
        self.titolo.setGeometry(QtCore.QRect(50, 14, 160, 32))
        self.titolo.setText("")
        self.titolo.setPixmap(QtGui.QPixmap("Immagini/titolo_casa_alfredo.png"))
        self.titolo.setScaledContents(True)
        self.titolo.setObjectName("titolo")
        self.scritta.setGeometry(QtCore.QRect(486, 320, 381, 91))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.scritta.setFont(font)
        self.scritta.setObjectName("scritta")
        self.registrati_button.setGeometry(QtCore.QRect(316, 482, 265, 61))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(33, 97, 171))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.registrati_button.setPalette(palette)
        font = QtGui.QFont()
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.registrati_button.setFont(font)
        self.registrati_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.registrati_button.setStyleSheet("QPushButton#registrati_button {\n"
                                             "background-color: rgb(33, 97, 171);\n"
                                             "border: 1px solid;\n"
                                             "border-color: rgb(13, 41, 73);\n"
                                             "border-radius: 10px;\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "}\n"
                                             "QPushButton#registrati_button:pressed {\n"
                                             "background-color: rgb(13, 41, 73);\n"
                                             "border-color: rgb(33, 97, 171);\n"
                                             "color: rgb(200, 200, 200);\n"
                                             "}")
        self.registrati_button.setObjectName("registrati_button")
        self.registrati_button.setFocusPolicy(QtCore.Qt.ClickFocus)
        self.banner.setGeometry(QtCore.QRect(0, 560, 901, 41))
        self.banner.setStyleSheet("background-color: rgb(13, 41, 73);")
        self.banner.setText("")
        self.banner.setObjectName("banner")
        self.torna_indietro.setGeometry(QtCore.QRect(730, 565, 171, 31))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.torna_indietro.setPalette(palette)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(True)
        font.setUnderline(True)
        font.setWeight(50)
        self.torna_indietro.setFont(font)
        self.torna_indietro.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.torna_indietro.setStyleSheet("QPushButton#torna_indietro {\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "}\n"
                                          "\n"
                                          "QPushButton#torna_indietro:pressed {\n"
                                          "background-color: rgb(0, 0, 0, 0);\n"
                                          "color: rgb(200, 200, 200);\n"
                                          "}")
        self.torna_indietro.setFlat(True)
        self.torna_indietro.setObjectName("torna_indietro")
        self.torna_indietro.setFocusPolicy(QtCore.Qt.ClickFocus)

        self.occhiello_barrato_button.setGeometry(QtCore.QRect(855, 225, 40, 40))
        self.occhiello_barrato_button.setStyleSheet("QPushButton#occhiello_barrato_button:pressed {\n"
                                                    "background-color: rgb(0, 0, 0, 0);\n"
                                                    "}")
        self.occhiello_barrato_button.setText("")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("Immagini/occhiello_barrato_password.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.occhiello_barrato_button.setIcon(icon1)
        self.occhiello_barrato_button.setIconSize(QtCore.QSize(30, 30))
        self.occhiello_barrato_button.setFlat(True)
        self.occhiello_barrato_button.setObjectName("occhiello_barrato_button")
        self.occhiello_barrato_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.occhiello_barrato_button.setFocusPolicy(QtCore.Qt.ClickFocus)

        self.occhiello_button.setGeometry(QtCore.QRect(855, 225, 40, 40))
        self.occhiello_button.setStyleSheet("QPushButton#occhiello_button:pressed {\n"
                                            "background-color: rgb(0, 0, 0, 0);\n"
                                            "}")
        self.occhiello_button.setText("")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("Immagini/occhiello_password.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.occhiello_button.setIcon(icon1)
        self.occhiello_button.setIconSize(QtCore.QSize(30, 30))
        self.occhiello_button.setFlat(True)
        self.occhiello_button.setObjectName("occhiello_button")
        self.occhiello_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.occhiello_button.setVisible(False)
        self.occhiello_button.setFocusPolicy(QtCore.Qt.ClickFocus)

        self.registrazione.setGeometry(QtCore.QRect(219, 17, 240, 32))
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.registrazione.setFont(font)
        self.registrazione.setStyleSheet("color: rgb(84, 149, 223);")
        self.registrazione.setObjectName("registrazione")
        self.dati_personali.setGeometry(QtCore.QRect(150, 70, 141, 21))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(True)
        font.setWeight(50)
        self.dati_personali.setFont(font)
        self.dati_personali.setObjectName("dati_personali")
        self.nome.setGeometry(QtCore.QRect(10, 120, 66, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.nome.setFont(font)
        self.nome.setObjectName("nome")
        self.cognome.setGeometry(QtCore.QRect(10, 165, 103, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.cognome.setFont(font)
        self.cognome.setObjectName("cognome")
        self.sesso.setGeometry(QtCore.QRect(10, 255, 67, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.sesso.setFont(font)
        self.sesso.setObjectName("sesso")
        self.citta_di_nascita.setGeometry(QtCore.QRect(10, 345, 160, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.citta_di_nascita.setFont(font)
        self.citta_di_nascita.setObjectName("citta_di_nascita")
        self.provincia.setGeometry(QtCore.QRect(10, 300, 101, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.provincia.setFont(font)
        self.provincia.setObjectName("provincia")
        self.data_di_nascita.setGeometry(QtCore.QRect(10, 210, 160, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.data_di_nascita.setFont(font)
        self.data_di_nascita.setObjectName("data_di_nascita")
        self.codice_fiscale.setGeometry(QtCore.QRect(10, 390, 145, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(False)
        font.setWeight(50)
        self.codice_fiscale.setFont(font)
        self.codice_fiscale.setObjectName("codice_fiscale")
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.sesso_combo_box.setFont(font)
        self.sesso_combo_box.setGeometry(QtCore.QRect(97, 250, 65, 36))
        self.sesso_combo_box.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                           "border: 1px solid;")
        self.sesso_combo_box.setObjectName("sesso_combo_box")

        # Combo-box sesso
        scegli_sesso = ["M", "F"]
        self.sesso_combo_box.addItems(scegli_sesso)
        pre_selected_m = "M"
        self.sesso_combo_box.setCurrentText(pre_selected_m)
        self.sesso_combo_box.setFocusPolicy(QtCore.Qt.ClickFocus)

        self.citta_di_nascita_combo_box.setFont(font)
        self.citta_di_nascita_combo_box.setGeometry(QtCore.QRect(170, 340, 250, 36))
        self.citta_di_nascita_combo_box.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                                      "border: 1px solid;")
        self.citta_di_nascita_combo_box.setEditable(False)
        self.citta_di_nascita_combo_box.setIconSize(QtCore.QSize(20, 20))
        self.citta_di_nascita_combo_box.setFrame(True)
        self.citta_di_nascita_combo_box.setModelColumn(0)
        self.citta_di_nascita_combo_box.setObjectName("citta_di_nascita_combo_box")
        self.citta_di_nascita_combo_box.setFocusPolicy(QtCore.Qt.ClickFocus)
        self.provincia_combo_box.setFont(font)
        self.provincia_combo_box.setGeometry(QtCore.QRect(169, 295, 251, 36))
        self.provincia_combo_box.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                               "border: 1px solid;")
        self.provincia_combo_box.setObjectName("provincia_combo_box")

        # Combo-box province e comuni, ti consente di scegliere prima la provincia e poi il comune
        self.provincia_combo_box.addItem('Seleziona una provincia')
        self.provincia_combo_box.addItems(self.province)
        self.provincia_combo_box.setFocusPolicy(QtCore.Qt.ClickFocus)

        self.nome_line_edit.setGeometry(QtCore.QRect(166, 115, 253, 36))
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.nome_line_edit.setFont(font)
        self.nome_line_edit.setStyleSheet("QLineEdit#nome_line_edit {\n"
                                          "border: 1px solid;\n"
                                          "padding-left: 8px;\n"
                                          "padding-right: 8px;\n"
                                          "}")
        self.nome_line_edit.setPlaceholderText("")
        self.nome_line_edit.setObjectName("nome_line_edit")
        self.cognome_line_edit.setGeometry(QtCore.QRect(166, 160, 253, 36))
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.cognome_line_edit.setFont(font)
        self.cognome_line_edit.setStyleSheet("QLineEdit#cognome_line_edit {\n"
                                             "border: 1px solid;\n"
                                             "padding-left: 8px;\n"
                                             "padding-right: 8px;\n"
                                             "}")
        self.cognome_line_edit.setPlaceholderText("")
        self.cognome_line_edit.setObjectName("cognome_line_edit")
        self.data_di_nascita_date_edit.setGeometry(QtCore.QRect(189, 205, 230, 36))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.data_di_nascita_date_edit.setFont(font)
        self.data_di_nascita_date_edit.setCalendarPopup(True)
        self.data_di_nascita_date_edit.setCurrentSectionIndex(0)
        self.data_di_nascita_date_edit.setObjectName("data_di_nascita_date_edit")
        self.codice_fiscale_line_edit.setGeometry(QtCore.QRect(175, 385, 244, 36))
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.codice_fiscale_line_edit.setFont(font)
        self.codice_fiscale_line_edit.setStyleSheet("QLineEdit#codice_fiscale_line_edit {\n"
                                                    "border: 1px solid;\n"
                                                    "padding-left: 8px;\n"
                                                    "padding-right: 8px;\n"
                                                    "}")
        self.codice_fiscale_line_edit.setPlaceholderText("")
        self.codice_fiscale_line_edit.setObjectName("codice_fiscale_line_edit")
        self.email_line_edit.setGeometry(QtCore.QRect(574, 115, 281, 36))
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.email_line_edit.setFont(font)
        self.email_line_edit.setStyleSheet("QLineEdit#email_line_edit {\n"
                                           "border: 1px solid;\n"
                                           "padding-left: 8px;\n"
                                           "padding-right: 8px;\n"
                                           "}")
        self.email_line_edit.setEchoMode(QtWidgets.QLineEdit.Normal)
        self.email_line_edit.setPlaceholderText("")
        self.email_line_edit.setObjectName("email_line_edit")
        self.password_line_edit.setGeometry(QtCore.QRect(605, 225, 250, 36))
        font = QtGui.QFont()
        font.setFamily("Verdana")
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.password_line_edit.setFont(font)
        self.password_line_edit.setStyleSheet("QLineEdit#password_line_edit {\n"
                                              "border: 1px solid;\n"
                                              "padding-left: 8px;\n"
                                              "padding-right: 8px;\n"
                                              "}")
        self.password_line_edit.setEchoMode(QtWidgets.QLineEdit.Password)
        self.password_line_edit.setPlaceholderText("")
        self.password_line_edit.setObjectName("password_line_edit")
        self.email.setGeometry(QtCore.QRect(495, 120, 71, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.email.setFont(font)
        self.email.setObjectName("email")
        self.password.setGeometry(QtCore.QRect(485, 230, 113, 25))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.password.setFont(font)
        self.password.setObjectName("password")
        self.dati_account.setGeometry(QtCore.QRect(620, 70, 141, 21))
        font = QtGui.QFont()
        font.setBold(False)
        font.setUnderline(True)
        font.setWeight(50)
        self.dati_account.setFont(font)
        self.dati_account.setObjectName("dati_account")

        self.retranslate_ui(registrazione_paziente)
        QtCore.QMetaObject.connectSlotsByName(registrazione_paziente)

    def retranslate_ui(self, registrazione_paziente):
        """
        Funzione che formatta il testo degli oggetti creati all'interno della view

        :param registrazione_paziente: Oggetto della view che rappresenta la view stessa
        :type registrazione_paziente: ViewRegistrazionePaziente
        """
        _translate = QtCore.QCoreApplication.translate
        registrazione_paziente.setWindowTitle(_translate("registrazione_paziente", "Clinica Casa Alfredo"))
        self.scritta.setText(_translate("registrazione_paziente", "*Attenzione! La password deve contenere \n"
                                                                  " almeno 8 caratteri, di cui almeno \n"
                                                                  " 1 numero, 1 lettera Maiuscola, \n"
                                                                  " 1 carattere speciale"))
        self.registrati_button.setText(_translate("registrazione_paziente", "Registrati ora"))
        self.torna_indietro.setText(_translate("registrazione_paziente", "Torna Indietro"))
        self.registrazione.setText(_translate("registrazione_paziente", "- REGISTRAZIONE"))
        self.dati_personali.setText(_translate("registrazione_paziente", "Dati Personali"))
        self.nome.setText(_translate("registrazione_paziente", "Nome:"))
        self.cognome.setText(_translate("registrazione_paziente", "Cognome:"))
        self.sesso.setText(_translate("registrazione_paziente", "Sesso:"))
        self.citta_di_nascita.setText(_translate("registrazione_paziente", "Città di nascita:"))
        self.provincia.setText(_translate("registrazione_paziente", "Provincia:"))
        self.data_di_nascita.setText(_translate("registrazione_paziente", "Data di nascita:"))
        self.codice_fiscale.setText(_translate("registrazione_paziente", "Codice fiscale:"))
        self.email.setText(_translate("registrazione_paziente", "E-mail:"))
        self.password.setText(_translate("registrazione_paziente", "*Password:"))
        self.dati_account.setText(_translate("registrazione_paziente", "Dati Account"))
